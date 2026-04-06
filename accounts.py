"""
accounts.py — CRUD router for X account management.

Secrets (api_key, api_secret, access_token, access_secret) are written to the
database but are NEVER returned in any API response.

Endpoints:
    POST   /accounts                  — add a new account
    GET    /accounts                  — list all accounts (no secrets)
    PATCH  /accounts/{id}/toggle      — flip is_active on/off
    DELETE /accounts/{id}             — remove account
    POST   /accounts/test-post/{id}   — post a test tweet from this account
    GET    /accounts/import/template  — download sample Excel template
    POST   /accounts/import           — bulk import from .xlsx or .csv
"""

import io

import pandas as pd
import openpyxl
from fastapi import APIRouter, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from database import get_connection

router = APIRouter(prefix="/accounts", tags=["Accounts"])


# ---------------------------------------------------------------------------
# Schemas
# ---------------------------------------------------------------------------

class AccountIn(BaseModel):
    username: str
    tone: str                        # formal | casual | aggressive | analytical | satirical
    persona_description: str
    api_key: str
    api_secret: str
    access_token: str
    access_secret: str


class AccountSummary(BaseModel):
    """Safe public view — no secrets."""
    id: int
    username: str
    tone: str | None
    persona_description: str | None
    is_active: int
    created_at: str


class AccountCreated(BaseModel):
    """Minimal response on creation — no secrets."""
    id: int
    username: str
    tone: str | None
    is_active: int


class ToggleResponse(BaseModel):
    id: int
    username: str
    is_active: int


class TestPostBody(BaseModel):
    text: str


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@router.post("", response_model=AccountCreated, status_code=201)
def create_account(body: AccountIn):
    """Store a new X account with its API credentials."""
    conn = get_connection()
    try:
        cursor = conn.execute(
            """
            INSERT INTO accounts
                (username, tone, persona_description,
                 api_key, api_secret, access_token, access_secret)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                body.username,
                body.tone,
                body.persona_description,
                body.api_key,
                body.api_secret,
                body.access_token,
                body.access_secret,
            ),
        )
        conn.commit()
        row = conn.execute(
            "SELECT id, username, tone, is_active FROM accounts WHERE id = ?",
            (cursor.lastrowid,),
        ).fetchone()
        return dict(row)
    except Exception as exc:
        # Unique constraint on username will surface here
        raise HTTPException(status_code=409, detail=str(exc)) from exc
    finally:
        conn.close()


@router.get("", response_model=list[AccountSummary])
def list_accounts():
    """Return all accounts — secrets excluded."""
    conn = get_connection()
    try:
        rows = conn.execute(
            """
            SELECT id, username, tone, persona_description, is_active, created_at
            FROM accounts
            ORDER BY created_at DESC
            """
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Bulk import — template download
# ---------------------------------------------------------------------------

# All columns the import endpoint expects (in display order)
IMPORT_COLUMNS = [
    "username",
    "tone",
    "persona_description",
    "api_key",
    "api_secret",
    "access_token",
    "access_secret",
]

VALID_TONES = {"formal", "casual", "aggressive", "analytical", "satirical"}


@router.get("/import/template")
def download_import_template():
    """
    Generate and return a sample .xlsx file with one header row and one
    example row so users know exactly what format to submit.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Accounts"

    # Header row
    ws.append(IMPORT_COLUMNS)

    # One example row — clearly placeholder values so nobody submits it by accident
    ws.append([
        "exampleuser",
        "formal",
        "Professional diplomatic tone",
        "your_api_key_here",
        "your_api_secret_here",
        "your_access_token_here",
        "your_access_secret_here",
    ])

    # Style header cells bold
    from openpyxl.styles import Font
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Set sensible column widths
    widths = [18, 12, 30, 28, 28, 28, 28]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # Write to an in-memory buffer and stream back
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="tweet_engine_accounts_template.xlsx"'
        },
    )


# ---------------------------------------------------------------------------
# Bulk import — file upload
# ---------------------------------------------------------------------------

@router.post("/import")
async def import_accounts(file: UploadFile = File(...)):
    """
    Accept a .xlsx or .csv file and bulk-insert valid account rows.

    Rules:
    - Required columns (case-insensitive): username, tone, persona_description,
      api_key, api_secret, access_token, access_secret
    - tone must be one of the five valid values (case-insensitive, stripped)
    - Rows with any empty required field are skipped
    - Rows whose username already exists in the DB are skipped
    - Returns a summary: {imported, skipped, errors}
    """
    filename = file.filename or ""
    if not (filename.endswith(".xlsx") or filename.endswith(".csv")):
        raise HTTPException(
            status_code=422,
            detail="Only .xlsx and .csv files are accepted",
        )

    content = await file.read()
    buf = io.BytesIO(content)

    # ── Parse file into a DataFrame ──────────────────────────────────────────
    try:
        if filename.endswith(".xlsx"):
            df = pd.read_excel(buf, engine="openpyxl")
        else:
            df = pd.read_csv(buf)
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Could not parse file: {exc}") from exc

    # Normalise column names to lowercase + stripped
    df.columns = [str(c).strip().lower() for c in df.columns]

    # Verify all required columns are present
    missing = [col for col in IMPORT_COLUMNS if col not in df.columns]
    if missing:
        raise HTTPException(
            status_code=422,
            detail=f"Missing required columns: {', '.join(missing)}",
        )

    # ── Fetch existing usernames to detect duplicates ────────────────────────
    conn = get_connection()
    try:
        existing = {
            row["username"]
            for row in conn.execute("SELECT username FROM accounts").fetchall()
        }
    finally:
        conn.close()

    # ── Process rows ─────────────────────────────────────────────────────────
    imported = 0
    skipped  = 0
    errors: list[str] = []

    for raw_idx, row in df.iterrows():
        row_num = int(raw_idx) + 2  # +2: 1-based + header row

        # Extract and clean each field
        values = {col: str(row[col]).strip() if pd.notna(row[col]) else "" for col in IMPORT_COLUMNS}

        # Skip rows where any required field is empty
        empty_fields = [col for col, val in values.items() if not val]
        if empty_fields:
            skipped += 1
            errors.append(f"row {row_num}: empty fields ({', '.join(empty_fields)}) — skipped")
            continue

        # Validate and normalise tone
        tone = values["tone"].lower()
        if tone not in VALID_TONES:
            skipped += 1
            errors.append(
                f"row {row_num}: invalid tone '{values['tone']}' "
                f"(must be one of: {', '.join(sorted(VALID_TONES))}) — skipped"
            )
            continue

        # Skip duplicate usernames
        username = values["username"]
        if username in existing:
            skipped += 1
            errors.append(f"row {row_num}: username '{username}' already exists — skipped")
            continue

        # Insert the row
        try:
            conn = get_connection()
            conn.execute(
                """
                INSERT INTO accounts
                    (username, tone, persona_description,
                     api_key, api_secret, access_token, access_secret)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    username,
                    tone,
                    values["persona_description"],
                    values["api_key"],
                    values["api_secret"],
                    values["access_token"],
                    values["access_secret"],
                ),
            )
            conn.commit()
            conn.close()
            existing.add(username)  # prevent duplicate within the same file
            imported += 1
        except Exception as exc:
            skipped += 1
            errors.append(f"row {row_num}: DB error — {exc}")

    return {"imported": imported, "skipped": skipped, "errors": errors}


# ---------------------------------------------------------------------------
# Parameterised routes — registered AFTER all literal paths so FastAPI
# doesn't accidentally shadow /import/template with /{account_id}
# ---------------------------------------------------------------------------

@router.post("/test-post/{account_id}")
def test_post(account_id: int, body: TestPostBody):
    """Post a test tweet from a specific account."""
    try:
        from poster import post_tweet
    except Exception as import_err:
        print(f"[accounts] Failed to import poster: {type(import_err).__name__}: {import_err}")
        raise HTTPException(status_code=500, detail=f"Import error: {import_err}")

    try:
        result = post_tweet(account_id, body.text)
        if not result["success"]:
            error_msg = result.get("error") or "Tweet posting failed (unknown error)"
            print(f"[accounts] Test post failed for account {account_id}: {error_msg}")
            raise HTTPException(status_code=502, detail=error_msg)
        return result
    except HTTPException:
        raise
    except Exception as e:
        print(f"[accounts] Unexpected error in test_post for account {account_id}: {type(e).__name__}: {e}")
        raise HTTPException(
            status_code=500,
            detail=str(e) or type(e).__name__,
        )


@router.patch("/{account_id}/toggle", response_model=ToggleResponse)
def toggle_active(account_id: int):
    """Flip is_active between 1 (active) and 0 (paused) for an account."""
    conn = get_connection()
    try:
        result = conn.execute(
            "UPDATE accounts SET is_active = 1 - is_active WHERE id = ?",
            (account_id,),
        )
        conn.commit()
        if result.rowcount == 0:
            raise HTTPException(status_code=404, detail="Account not found")
        row = conn.execute(
            "SELECT id, username, is_active FROM accounts WHERE id = ?",
            (account_id,),
        ).fetchone()
        return dict(row)
    finally:
        conn.close()


@router.delete("/{account_id}", status_code=204)
def delete_account(account_id: int):
    """Delete an account (cascades to tweet_queue and post_history)."""
    conn = get_connection()
    try:
        result = conn.execute(
            "DELETE FROM accounts WHERE id = ?", (account_id,)
        )
        conn.commit()
        if result.rowcount == 0:
            raise HTTPException(status_code=404, detail="Account not found")
    finally:
        conn.close()
